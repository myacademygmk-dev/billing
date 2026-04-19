from __future__ import annotations

import json
import re
import uuid
from calendar import month_name
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

from fastapi import APIRouter, Depends, Form, HTTPException, Query, Response, UploadFile
from fastapi import File as FastAPIFile
from openpyxl import load_workbook
from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session

from app.api.deps import get_current_user, require_admin_user
from app.core.database import get_db
from app.models.payment import Payment
from app.models.student import Student
from app.models.student_fee import StudentFee
from app.models.user import User
from app.models.enums import StudentStatus
from app.schemas.students import (
    StudentCreate,
    StudentBillingMonthStatus,
    StudentBillingOverviewRead,
    StudentListItem,
    StudentBalanceRead,
    StudentFeeRead,
    StudentFeeUpdate,
    StudentImportMapping,
    StudentImportPreviewRead,
    StudentImportResult,
    StudentRead,
    StudentUpdate,
)
from app.services.billing import billing_list_snapshot, get_student_billing_overview, month_label, pending_amount


router = APIRouter()

_HEADER_RE = re.compile(r"[^a-z0-9]+")
_BATCH_RE = re.compile(r"^\d{4}\s*-\s*\d{4}$")
_REQUIRED_IMPORT_FIELDS = [
    "serial_no",
    "student_code",
    "name",
    "class_name",
    "expected_fee",
    "payment_period",
    "joined_date",
    "billing_start_period",
    "billing_end_period",
]


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    s = str(value).strip().lower()
    s = _HEADER_RE.sub("", s)
    return s


def _find_col(headers: dict[str, int], keys: list[str]) -> int | None:
    for key in keys:
        idx = headers.get(key)
        if idx is not None:
            return idx
    return None


def _find_header_name(headers: dict[str, int], keys: list[str], ordered_headers: list[str]) -> str | None:
    idx = _find_col(headers, keys)
    if idx is None:
        return None
    return ordered_headers[idx]


def _to_student_code(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _to_int(value: object) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        raise ValueError("invalid integer")
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if not value.is_integer():
            raise ValueError("invalid integer")
        return int(value)
    s = str(value).strip()
    if s == "":
        return None
    return int(s)


def _to_decimal(value: object) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    s = str(value).strip()
    if s == "":
        return None
    return Decimal(s)


def _to_date(value: object) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if s == "":
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError("invalid date")


def _preview_cell_value(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _read_workbook(file_bytes: bytes):
    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception:
        raise HTTPException(status_code=422, detail="Invalid Excel file")

    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        raise HTTPException(status_code=422, detail="Excel file is empty")
    ordered_headers = ["" if value is None else str(value).strip() for value in header_row]
    normalized_headers: dict[str, int] = {}
    for idx, value in enumerate(header_row):
        key = _normalize_header(value)
        if key and key not in normalized_headers:
            normalized_headers[key] = idx
    return ws, rows, ordered_headers, normalized_headers


def _default_import_mapping(headers: dict[str, int], ordered_headers: list[str]) -> StudentImportMapping:
    return StudentImportMapping(
        serial_no=_find_header_name(headers, ["sno", "serialno", "serialnumber", "slno"], ordered_headers),
        student_code=_find_header_name(
            headers,
            ["studentcode", "studentid", "rollno", "rollnumber", "roll", "id", "roll_no", "rno"],
            ordered_headers,
        ),
        name=_find_header_name(headers, ["name", "studentname"], ordered_headers),
        class_name=_find_header_name(headers, ["classname", "class", "std", "standard"], ordered_headers),
        expected_fee=_find_header_name(headers, ["expectedfeeamount", "expectedfee", "fee", "fees"], ordered_headers),
        payment_period=_find_header_name(headers, ["paymentperiod", "period", "cycle", "paymentcycle"], ordered_headers),
        joined_date=_find_header_name(headers, ["joineddate", "datejoined", "joindate", "admissiondate"], ordered_headers),
        billing_start_period=_find_header_name(headers, ["start", "startmonth", "startperiod", "periodstart"], ordered_headers),
        billing_end_period=_find_header_name(headers, ["end", "endmonth", "endperiod", "periodend"], ordered_headers),
    )


def _header_index_by_label(ordered_headers: list[str]) -> dict[str, int]:
    labels: dict[str, int] = {}
    for idx, header in enumerate(ordered_headers):
        label = header.strip()
        if label and label not in labels:
            labels[label] = idx
    return labels


def _split_class_and_section(value: str | None) -> tuple[str | None, str | None]:
    if value is None:
        return None, None
    text = value.strip()
    if text == "":
        return None, None
    if "-" not in text:
        return text, None
    left, right = text.split("-", 1)
    left = left.strip()
    right = right.strip()
    if left and right and right.replace(" ", "").isalpha():
        return left, right
    return text, None


def _validate_batch(batch: str) -> str:
    normalized = batch.strip()
    if not _BATCH_RE.match(normalized):
        raise HTTPException(status_code=422, detail='batch must be in the format "YYYY-YYYY"')
    return normalized.replace(" ", "")


def _parse_batch_start_month(value: str) -> int:
    normalized = value.strip().lower()
    if normalized.isdigit():
        month_number = int(normalized)
        if 1 <= month_number <= 12:
            return month_number
    for month_number in range(1, 13):
        full_name = month_name[month_number].lower()
        if normalized in {full_name, full_name[:3]}:
            return month_number
    raise HTTPException(status_code=422, detail="batch_start_month must be a month name like May or a number 1-12")


def _parse_month_value(value: object, field_name: str) -> int:
    if value is None:
        raise ValueError(f"{field_name} is required")
    if isinstance(value, (datetime, date)):
        return value.month
    text = str(value).strip()
    if text == "":
        raise ValueError(f"{field_name} is required")
    if text.isdigit():
        month_number = int(text)
        if 1 <= month_number <= 12:
            return month_number
    normalized = text.lower()
    for month_number in range(1, 13):
        full_name = month_name[month_number].lower()
        if normalized in {full_name, full_name[:3]}:
            return month_number
    raise ValueError(f"{field_name} must be a month name like Jun or a number 1-12")


@router.post("/import/preview", response_model=StudentImportPreviewRead)
async def preview_students_import(
    _: User = Depends(require_admin_user),
    file: UploadFile = FastAPIFile(...),
) -> StudentImportPreviewRead:
    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(status_code=415, detail="Only .xlsx files are supported")

    data = await file.read()
    _, rows, ordered_headers, headers = _read_workbook(data)
    sample_rows: list[dict[str, str | None]] = []
    for row in rows:
        if not any(cell is not None and str(cell).strip() != "" for cell in row):
            continue
        row_data: dict[str, str | None] = {}
        for idx, header in enumerate(ordered_headers):
            key = header or f"Column {idx + 1}"
            value = row[idx] if idx < len(row) else None
            row_data[key] = _preview_cell_value(value)
        sample_rows.append(row_data)
        if len(sample_rows) >= 5:
            break

    return StudentImportPreviewRead(
        headers=[header or f"Column {idx + 1}" for idx, header in enumerate(ordered_headers)],
        suggested_mapping=_default_import_mapping(headers, ordered_headers),
        sample_rows=sample_rows,
        required_fields=_REQUIRED_IMPORT_FIELDS,
    )


@router.post("/import", response_model=StudentImportResult)
async def import_students_from_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_user),
    file: UploadFile = FastAPIFile(...),
    mapping_json: str | None = Form(default=None),
    batch: str = Form(...),
    batch_start_month: str | None = Form(default=None),
    mode: str = Query("upsert", pattern="^(upsert|create_only)$"),
    atomic: bool = Query(True),
) -> StudentImportResult:
    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(status_code=415, detail="Only .xlsx files are supported")

    data = await file.read()
    _, rows, ordered_headers, headers = _read_workbook(data)
    normalized_batch = _validate_batch(batch)
    normalized_batch_start_month = _parse_batch_start_month(batch_start_month) if batch_start_month else None
    if mapping_json:
        try:
            mapping = StudentImportMapping.model_validate(json.loads(mapping_json))
        except Exception:
            raise HTTPException(status_code=422, detail="Invalid mapping_json payload")
    else:
        mapping = _default_import_mapping(headers, ordered_headers)

    header_lookup = _header_index_by_label(ordered_headers)
    missing = [field for field in _REQUIRED_IMPORT_FIELDS if not getattr(mapping, field)]
    if missing:
        raise HTTPException(status_code=422, detail=f"Missing required mapping for: {', '.join(missing)}")

    def mapped_col(header_name: str | None) -> int | None:
        if not header_name:
            return None
        idx = header_lookup.get(header_name)
        if idx is None:
            raise HTTPException(status_code=422, detail=f"Mapped column '{header_name}' was not found in the file")
        return idx

    serial_no_col = mapped_col(mapping.serial_no)
    student_code_col = mapped_col(mapping.student_code)
    name_col = mapped_col(mapping.name)
    class_col = mapped_col(mapping.class_name)
    fee_col = mapped_col(mapping.expected_fee)
    payment_period_col = mapped_col(mapping.payment_period)
    joined_date_col = mapped_col(mapping.joined_date)
    billing_start_col = mapped_col(mapping.billing_start_period)
    billing_end_col = mapped_col(mapping.billing_end_period)

    created = 0
    updated = 0
    fee_updated = 0
    errors: list[dict] = []

    # We process and write in one pass; if atomic=True we roll back on any error.
    try:
        for row_index, row in enumerate(rows, start=2):
            def cell(col: int | None) -> object:
                if col is None:
                    return None
                return row[col] if col < len(row) else None

            try:
                serial_no = _to_int(cell(serial_no_col))
            except ValueError:
                errors.append({"row": row_index, "error": "serial_no must be an integer"})
                continue
            student_code = _to_student_code(cell(student_code_col))
            name_val = (cell(name_col) if name_col is not None else None)
            name = "" if name_val is None else str(name_val).strip()
            class_raw = None
            if class_col is not None:
                v = cell(class_col)
                class_raw = None if v is None or str(v).strip() == "" else str(v).strip()
            class_name, section = _split_class_and_section(class_raw)
            payment_period_val = cell(payment_period_col)
            payment_period = "" if payment_period_val is None else str(payment_period_val).strip()
            billing_start_cell = cell(billing_start_col)
            billing_end_cell = cell(billing_end_col)

            fee_cell = cell(fee_col) if fee_col is not None else None
            fee_cell_empty = fee_cell is None or str(fee_cell).strip() == ""
            joined_date_cell = cell(joined_date_col)
            joined_date_empty = joined_date_cell is None or str(joined_date_cell).strip() == ""
            billing_start_empty = billing_start_cell is None or str(billing_start_cell).strip() == ""
            billing_end_empty = billing_end_cell is None or str(billing_end_cell).strip() == ""
            if student_code == "" and name == "" and (class_name is None) and fee_cell_empty and payment_period == "" and joined_date_empty and billing_start_empty and billing_end_empty:
                continue  # empty-ish row

            if student_code == "":
                errors.append({"row": row_index, "error": "student_code/rollno is required"})
                continue
            if name == "":
                errors.append({"row": row_index, "error": "name is required"})
                continue
            if class_name is None:
                errors.append({"row": row_index, "error": "class is required"})
                continue
            if payment_period == "":
                errors.append({"row": row_index, "error": "payment_period is required"})
                continue
            try:
                billing_start_month = _parse_month_value(billing_start_cell, "billing_start_period")
            except ValueError as exc:
                errors.append({"row": row_index, "error": str(exc)})
                continue
            try:
                billing_end_month = _parse_month_value(billing_end_cell, "billing_end_period")
            except ValueError as exc:
                errors.append({"row": row_index, "error": str(exc)})
                continue

            try:
                joined_date = _to_date(joined_date_cell)
            except ValueError:
                errors.append({"row": row_index, "error": "joined_date must be a valid date"})
                continue
            if joined_date is None:
                errors.append({"row": row_index, "error": "joined_date is required"})
                continue

            try:
                fee = _to_decimal(cell(fee_col)) if fee_col is not None else None
            except (InvalidOperation, ValueError):
                errors.append({"row": row_index, "error": "fees must be a number"})
                continue
            if fee is None:
                fee = Decimal("0")
            if fee < 0:
                errors.append({"row": row_index, "error": "fees must be >= 0"})
                continue

            student = db.execute(select(Student).where(Student.student_code == student_code)).scalar_one_or_none()
            if student:
                if mode == "create_only":
                    errors.append({"row": row_index, "error": f"student_code '{student_code}' already exists"})
                    continue
                student.serial_no = serial_no
                student.name = name
                student.class_name = class_name
                student.section = section
                student.payment_period = payment_period
                student.joined_date = joined_date
                student.batch = normalized_batch
                student.batch_start_month = normalized_batch_start_month
                student.billing_start_month = billing_start_month
                student.billing_end_month = billing_end_month
                updated += 1
            else:
                student = Student(
                    serial_no=serial_no,
                    student_code=student_code,
                    name=name,
                    class_name=class_name,
                    section=section,
                    payment_period=payment_period,
                    joined_date=joined_date,
                    batch=normalized_batch,
                    batch_start_month=normalized_batch_start_month,
                    billing_start_month=billing_start_month,
                    billing_end_month=billing_end_month,
                )
                db.add(student)
                db.flush()
                created += 1

            fee_row = db.get(StudentFee, student.id)
            if not fee_row:
                fee_row = StudentFee(student_id=student.id)
                db.add(fee_row)
            fee_row.expected_fee_amount = fee
            fee_row.last_fee_updated_at = datetime.now(UTC)
            fee_row.last_fee_updated_by = current_user.id
            fee_updated += 1

        if errors and atomic:
            db.rollback()
            raise HTTPException(status_code=422, detail={"errors": errors})

        db.commit()
    except HTTPException:
        raise
    except Exception:
        db.rollback()
        raise

    return StudentImportResult(
        created=created,
        updated=updated,
        fee_updated=fee_updated,
        errors=errors,
    )


@router.get("", response_model=dict)
def list_students(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
    search: str | None = None,
    class_code: str | None = None,
    status: StudentStatus | None = None,
    class_name: str | None = None,
    section: str | None = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
) -> dict:
    stmt = select(Student)

    if search:
        s = f"%{search.lower()}%"
        stmt = stmt.where(
            or_(func.lower(Student.student_code).like(s), func.lower(Student.name).like(s))
        )
    if class_code:
        class_code_normalized = class_code.strip()
        stmt = stmt.where(Student.student_code.like(f"{class_code_normalized}%"))
    if status is not None:
        stmt = stmt.where(Student.status == status)
    if class_name:
        stmt = stmt.where(Student.class_name == class_name)
    if section:
        stmt = stmt.where(Student.section == section)

    total = db.execute(select(func.count()).select_from(stmt.subquery())).scalar_one()
    items = (
        db.execute(
            stmt.order_by(Student.student_code)
            .offset((page - 1) * page_size)
            .limit(page_size)
        )
        .scalars()
        .all()
    )
    return {"items": [StudentRead.model_validate(s) for s in items], "total": total}


@router.get("/balances", response_model=dict)
def list_student_balances(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
    search: str | None = None,
    class_code: str | None = None,
    status: StudentStatus | None = None,
    class_name: str | None = None,
    section: str | None = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
) -> dict:
    stmt = select(Student)

    if search:
        s = f"%{search.lower()}%"
        stmt = stmt.where(
            or_(func.lower(Student.student_code).like(s), func.lower(Student.name).like(s))
        )
    if class_code:
        class_code_normalized = class_code.strip()
        stmt = stmt.where(Student.student_code.like(f"{class_code_normalized}%"))
    if status is not None:
        stmt = stmt.where(Student.status == status)
    if class_name:
        stmt = stmt.where(Student.class_name == class_name)
    if section:
        stmt = stmt.where(Student.section == section)

    total = db.execute(select(func.count()).select_from(stmt.subquery())).scalar_one()
    rows = (
        db.execute(
            stmt.order_by(Student.student_code)
            .offset((page - 1) * page_size)
            .limit(page_size)
        )
        .scalars()
        .all()
    )
    items = []
    for student in rows:
        overview = get_student_billing_overview(db, student)
        snapshot = billing_list_snapshot(overview)
        paid_total = db.execute(
            select(func.coalesce(func.sum(Payment.amount), 0)).where(Payment.student_id == student.id)
        ).scalar_one()
        items.append(
            StudentListItem(
                id=student.id,
                serial_no=student.serial_no,
                student_code=student.student_code,
                name=student.name,
                class_name=student.class_name,
                section=student.section,
                payment_period=student.payment_period,
                joined_date=student.joined_date,
                batch=student.batch,
                status=student.status,
                billing_start_month=student.billing_start_month,
                billing_end_month=student.billing_end_month,
                expected_fee=overview.monthly_fee,
                paid_total=paid_total,
                pending=pending_amount(overview),
                last_paid_label=snapshot.last_paid_label,
                next_due_label=snapshot.next_due_label,
                next_due_state=snapshot.next_due_state,
            )
        )
    return {"items": items, "total": total}


@router.post("", response_model=StudentRead, status_code=201)
def create_student(
    payload: StudentCreate,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
) -> StudentRead:
    existing = db.execute(select(Student).where(Student.student_code == payload.student_code)).scalar_one_or_none()
    if existing:
        raise HTTPException(status_code=409, detail="student_code already exists")

    student = Student(
        serial_no=payload.serial_no,
        student_code=payload.student_code,
        name=payload.name,
        class_name=payload.class_name,
        section=payload.section,
        payment_period=payload.payment_period,
        joined_date=payload.joined_date,
        batch=payload.batch,
        batch_start_month=payload.batch_start_month,
        billing_start_month=payload.billing_start_month,
        billing_end_month=payload.billing_end_month,
    )
    student.fee = StudentFee(expected_fee_amount=0)
    db.add(student)
    db.commit()
    db.refresh(student)
    return StudentRead.model_validate(student)


@router.get("/{student_id}", response_model=StudentRead)
def get_student(
    student_id: uuid.UUID,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
) -> StudentRead:
    student = db.get(Student, student_id)
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    return StudentRead.model_validate(student)


@router.get("/{student_id}/balance", response_model=StudentBalanceRead)
def get_student_balance(
    student_id: uuid.UUID,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
) -> StudentBalanceRead:
    student = db.get(Student, student_id)
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    overview = get_student_billing_overview(db, student)
    paid_total = db.execute(
        select(func.coalesce(func.sum(Payment.amount), 0)).where(Payment.student_id == student_id)
    ).scalar_one()
    return StudentBalanceRead(
        student_id=student.id,
        student_code=student.student_code,
        name=student.name,
        expected_fee=overview.monthly_fee,
        paid_total=paid_total,
        pending=pending_amount(overview),
    )


@router.patch("/{student_id}", response_model=StudentRead)
def update_student(
    student_id: uuid.UUID,
    payload: StudentUpdate,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
) -> StudentRead:
    student = db.get(Student, student_id)
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    data = payload.model_dump(exclude_unset=True)
    for key, value in data.items():
        setattr(student, key, value)

    db.commit()
    db.refresh(student)
    return StudentRead.model_validate(student)


@router.patch("/{student_id}/fee", response_model=StudentFeeRead)
def update_student_fee(
    student_id: uuid.UUID,
    payload: StudentFeeUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> StudentFeeRead:
    student = db.get(Student, student_id)
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    fee = db.get(StudentFee, student_id)
    if not fee:
        fee = StudentFee(student_id=student_id)
        db.add(fee)

    fee.expected_fee_amount = payload.expected_fee_amount
    fee.last_fee_updated_at = datetime.now(UTC)
    fee.last_fee_updated_by = current_user.id

    db.commit()
    db.refresh(fee)
    return StudentFeeRead.model_validate(fee)


@router.get("/{student_id}/fee", response_model=StudentFeeRead)
def get_student_fee(
    student_id: uuid.UUID,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
) -> StudentFeeRead:
    fee = db.get(StudentFee, student_id)
    if not fee:
        fee = StudentFee(student_id=student_id)
        db.add(fee)
        db.commit()
        db.refresh(fee)
    return StudentFeeRead.model_validate(fee)


@router.get("/{student_id}/billing-overview", response_model=StudentBillingOverviewRead)
def get_student_billing_overview_route(
    student_id: uuid.UUID,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
) -> StudentBillingOverviewRead:
    student = db.get(Student, student_id)
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    overview = get_student_billing_overview(db, student)
    return StudentBillingOverviewRead(
        student_id=student.id,
        monthly_fee=overview.monthly_fee,
        cycle_label=overview.cycle_label,
        cycle_months=overview.cycle_months,
        payable_amount=overview.payable_amount,
        batch=overview.batch,
        batch_start_month=overview.batch_start_month,
        batch_start_label=overview.batch_start_label,
        batch_end_label=overview.batch_end_label,
        next_unpaid_month=overview.next_unpaid_month,
        next_unpaid_label=month_label(overview.next_unpaid_month),
        pending_months=[StudentBillingMonthStatus.model_validate(item) for item in overview.pending_months],
        months=[StudentBillingMonthStatus.model_validate(item) for item in overview.months],
    )


@router.delete("/{student_id}", status_code=204, response_class=Response)
def hard_delete_student(
    student_id: uuid.UUID,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin_user),
) -> Response:
    student = db.get(Student, student_id)
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    if student.status != StudentStatus.inactive:
        raise HTTPException(status_code=409, detail="Only inactive students can be permanently deleted")
    has_payments = db.execute(select(func.count()).select_from(Payment).where(Payment.student_id == student_id)).scalar_one()
    if has_payments:
        raise HTTPException(status_code=409, detail="Cannot delete inactive student with payment history")
    db.delete(student)
    db.commit()
    return Response(status_code=204)
